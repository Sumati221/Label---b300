"""
B300 Label Generator v6

Pipeline:
1. PyMuPDF (fitz): rasterize PDF page + extract text spans with bounding boxes
2. OpenCV: multi-scale template matching of each PNG/SVG symbol against the
   rendered PDF (transparency-aware binary masks)
3. Output: normalized coords for symbol placement + editable text fields
4. No hard-coded positions — everything detected from the selected PDF
"""
import os
import re
import io
import base64
import logging
import time
import tempfile
import uuid
from pathlib import Path
from typing import List, Dict, Optional
from html.parser import HTMLParser

import numpy as np
from PIL import Image, ImageDraw, ImageFont

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import fitz  # PyMuPDF
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False

try:
    import cv2
    HAS_CV2 = True
except ImportError:
    HAS_CV2 = False

try:
    import cairosvg
    HAS_CAIRO = True
except (ImportError, OSError):
    HAS_CAIRO = False

from fastapi import FastAPI, File, HTTPException, UploadFile
from pydantic import BaseModel, Field
from fastapi.responses import HTMLResponse, Response, JSONResponse, StreamingResponse
import json

logging.basicConfig(level=logging.INFO)
log = logging.getLogger("label-v6")


class NumpyEncoder(json.JSONEncoder):
    """Handle numpy int32/float32 from OpenCV in JSON responses."""
    def default(self, obj):
        if isinstance(obj, (np.integer,)):
            return int(obj)
        if isinstance(obj, (np.floating,)):
            return float(obj)
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        return super().default(obj)


def sanitize(obj):
    """Recursively convert numpy types to native Python for JSON.
    Also handles inf/nan which are not valid JSON."""
    if isinstance(obj, dict):
        return {k: sanitize(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [sanitize(v) for v in obj]
    if isinstance(obj, (np.integer,)):
        return int(obj)
    if isinstance(obj, (np.floating,)):
        v = float(obj)
        if v != v:  # nan
            return 0.0
        if v == float('inf') or v == float('-inf'):
            return 1.0
        return v
    if isinstance(obj, float):
        if obj != obj:  # nan
            return 0.0
        if obj == float('inf') or obj == float('-inf'):
            return 1.0
        return obj
    if isinstance(obj, np.ndarray):
        return obj.tolist()
    return obj


def unwrap_ai_extract_values(value):
    """Flatten ai_extract v2.1 scalar wrappers (for example ``{"value": "x"}``)."""
    if isinstance(value, dict):
        if set(value) == {"value"}:
            return unwrap_ai_extract_values(value["value"])
        return {key: unwrap_ai_extract_values(item) for key, item in value.items()}
    if isinstance(value, list):
        return [unwrap_ai_extract_values(item) for item in value]
    return value


class CountryHtmlParser(HTMLParser):
    """Small, safe rich-text subset used by the downloadable label renderer."""
    def __init__(self):
        super().__init__()
        self.lines = [[]]
        self.styles = [{"bold": False, "italic": False, "size": None}]

    def _style_from_attrs(self, attrs):
        style = dict(self.styles[-1])
        css = dict(attrs).get("style", "")
        for item in css.split(";"):
            key, _, value = item.partition(":")
            key, value = key.strip().lower(), value.strip().lower()
            if key == "font-weight" and (value == "bold" or value.isdigit() and int(value) >= 600):
                style["bold"] = True
            elif key == "font-style" and value == "italic":
                style["italic"] = True
            elif key == "font-size" and re.fullmatch(r"[0-9.]+(?:pt|px)", value):
                style["size"] = value
        return style

    def handle_starttag(self, tag, attrs):
        tag = tag.lower()
        if tag == "br":
            self.lines.append([])
            return
        if tag in {"div", "p"} and self.lines[-1]:
            self.lines.append([])
        style = self._style_from_attrs(attrs)
        if tag in {"b", "strong"}:
            style["bold"] = True
        if tag in {"i", "em"}:
            style["italic"] = True
        self.styles.append(style)

    def handle_endtag(self, tag):
        if tag.lower() in {"div", "p"} and self.lines[-1]:
            self.lines.append([])
        if len(self.styles) > 1:
            self.styles.pop()

    def handle_data(self, data):
        if not data:
            return
        for index, fragment in enumerate(data.split("\n")):
            if fragment:
                self.lines[-1].append((fragment, dict(self.styles[-1])))
            if index < len(data.split("\n")) - 1:
                self.lines.append([])


def _export_font(size_px: float, bold: bool, italic: bool):
    suffix = "BoldOblique" if bold and italic else "Bold" if bold else "Oblique" if italic else ""
    filename = f"DejaVuSans{('-' + suffix) if suffix else ''}.ttf"
    for directory in ("/usr/share/fonts/truetype/dejavu", "/usr/share/fonts/dejavu"):
        path = Path(directory) / filename
        if path.exists():
            return ImageFont.truetype(str(path), max(6, round(size_px)))
    return ImageFont.load_default()


def render_country_html(draw, region: Dict, html: str, dpi: int, width: int, height: int):
    """Render the editor's safe rich-text subset into its physical label area."""
    x0 = int(float(region.get("x", 0)) * width)
    y0 = int(float(region.get("y", 0)) * height)
    x1 = int((float(region.get("x", 0)) + float(region.get("w", 1))) * width)
    y1 = int((float(region.get("y", 0)) + float(region.get("h", 1))) * height)
    draw.rectangle((x0, y0, x1, y1), fill="white")
    parser = CountryHtmlParser()
    parser.feed(html or "")
    default_size = max(8, float(region.get("font_size", 2.8)) / 25.4 * dpi)
    line_height_factor = float(region.get("line_height", 1.18))
    cursor_y, padding, available_width = y0 + max(3, dpi // 150), max(3, dpi // 150), max(1, x1 - x0 - 2 * max(3, dpi // 150))

    for line in parser.lines:
        if cursor_y >= y1 - padding:
            break
        cursor_x, line_height = x0 + padding, default_size * line_height_factor
        for text, style in line:
            chunks = re.split(r"(\s+)", text)
            raw_size = style.get("size")
            if raw_size and raw_size.endswith("pt"):
                font_size = float(raw_size[:-2]) / 72 * dpi
            elif raw_size and raw_size.endswith("px"):
                font_size = float(raw_size[:-2]) / 96 * dpi
            else:
                font_size = default_size
            font = _export_font(font_size, style.get("bold", False), style.get("italic", False))
            line_height = max(line_height, font_size * line_height_factor)
            for chunk in chunks:
                if not chunk:
                    continue
                chunk_width = draw.textlength(chunk, font=font)
                if not chunk.isspace() and cursor_x > x0 + padding and cursor_x + chunk_width > x0 + padding + available_width:
                    cursor_x = x0 + padding
                    cursor_y += int(line_height)
                    if cursor_y + line_height > y1 - padding:
                        return
                if cursor_y + line_height > y1 - padding:
                    return
                draw.text((cursor_x, cursor_y), chunk, fill="black", font=font)
                cursor_x += chunk_width
        cursor_y += int(line_height)


def render_thai_symbol_text(draw, region: Dict, text: str, dpi: int, width: int, height: int):
    """Replace only the variable Thai FDA notification number inside its frame."""
    value = (text or "").strip()
    if not value:
        return
    x0 = int(float(region["x"]) * width)
    y0 = int(float(region["y"]) * height)
    x1 = int((float(region["x"]) + float(region["w"])) * width)
    y1 = int((float(region["y"]) + float(region["h"])) * height)
    padding = max(2, dpi // 200)
    draw.rectangle((x0, y0, x1, y1), fill="white")
    max_width = max(1, x1 - x0 - 2 * padding)
    font_size = max(6, float(region.get("font_size", 1.7)) / 25.4 * dpi)
    font = _export_font(font_size, False, False)
    while font_size > 6 and draw.textlength(value, font=font) > max_width:
        font_size -= 1
        font = _export_font(font_size, False, False)
    text_y = y0 + max(0, ((y1 - y0) - int(font_size * 1.18)) // 2)
    draw.text((x0 + padding, text_y), value, fill="black", font=font)


app = FastAPI(title="B300 Label Generator v6", docs_url="/docs")

_APP = Path(__file__).parent
_SYMBOLS = _APP / "data" / "symbols"
_SYMBOL_ASSETS = _SYMBOLS / "assets"
_LABEL_GUIDES = _APP / "data" / "label_guides"
_CDLM = _APP / "data" / "cdlm"
_DRAWINGS = _APP / "data" / "drawings"
RENDER_DPI = 300  # DPI for PDF rasterization during matching
# Component IoU is deliberately conservative. A false positive must never add
# an unrelated regulatory mark to a label simply because the asset library grew.
MATCH_THRESHOLD = 0.22
MATCH_UNIQUENESS_MARGIN = 0.035
LABEL_EDGE_MARGIN_MM = 1.0  # Keep all placed symbols clear of the blank label border
_layout_manifest_cache = None
_drawing_catalog_cache = None
MAX_UPLOAD_BYTES = 25 * 1024 * 1024  # Uploads are processed temporarily, never retained.

# This schema turns a human edit request into a bounded operation.  The client
# still applies the edit visibly and the human decides whether to keep it.
AI_EDIT_SCHEMA = {
    "operation": {"type": "string", "description": "One of move_symbol, resize_symbol, replace_text, or unsupported."},
    "target": {"type": "string", "description": "Symbol name/code or the label text to edit."},
    "direction": {"type": "string", "description": "left, right, up, down, increase, decrease, or none."},
    "amount": {"type": "number", "description": "Numeric movement or resize amount, otherwise 0."},
    "unit": {"type": "string", "description": "mm, percent, or none."},
    "replacement_text": {"type": "string", "description": "Replacement label text when operation is replace_text; otherwise empty."},
}


class EditRequest(BaseModel):
    command: str = Field(min_length=1, max_length=2000)


class LabelExportRequest(BaseModel):
    label_id: str = "label"
    label_image: str
    w_mm: float = Field(gt=0, le=200)
    h_mm: float = Field(gt=0, le=200)
    dpi: int = Field(ge=300, le=600)
    symbols: List[Dict] = []
    symbol_images: Dict[str, str] = {}
    reference_graphics: List[Dict] = []
    country_region: Optional[Dict] = None
    country_html: str = ""
    thai_symbol_region: Optional[Dict] = None
    thai_symbol_text: str = Field(default="", max_length=100)
    thai_symbol_y: Optional[float] = Field(default=None, ge=0, le=1)
    layout_manifest: Optional[Dict] = None


class DrawingExportRequest(LabelExportRequest):
    """A generated label placed on its approved blank-label drawing."""
    drawing_part_number: str = Field(min_length=12, max_length=20)

# Extracted from the controlled specification documents. These values are kept
# with the app so label generation does not depend on a live AI/SQL request.
SYMBOL_SPECIFICATIONS = {
    "100025": {
        "symbol_id": "100025",
        "symbol_name": "INMETRO Brazilian National Institute of Metrology symbol",
        "reference": "INMETRO Ordinance 54/2016",
        "required_width_mm": 20.0,
        "minimum_size_mm": None,
        "size_specification": "Width 20 mm.",
        "applicable_for": "Product",
        "color_specification": None,
        "usage_notes": "Required symbol; no required combination with other symbols.",
        "source_document": "LS_100025_SPEC.docx",
    },
    "100183": {
        "symbol_id": "100183",
        "symbol_name": "Thai FDA logo",
        "reference": "Thailand Ministry of Public Health, Volume 137, Special Section 260, November 5, 2020",
        "required_width_mm": None,
        "minimum_size_mm": 5.0,
        "size_specification": "Minimum size 5 mm; text within the symbol must remain legible.",
        "applicable_for": "Product and packaging of medical devices sold in Thailand",
        "color_specification": "Not specified internally or by regulation.",
        "usage_notes": "Must contain the license number, detailed notification number, or notification receipt number in Arabic numerals within the Thai FDA logo frame.",
        "source_document": "LS-100183.docx",
    },
}

_cache = None
_cache_t = 0
_cdlm_cache = None
_specification_cache = None
_uploaded_cdlm_sessions: Dict[str, Dict] = {}

COUNTRY_ALIASES = {
    "Brazil": ("brazil", "brasil"),
    "Thailand": ("thailand", "thai"),
    "Philippines": ("philippines", "philippine"),
    "Mexico": ("mexico", "méxico", "mexican"),
    "Malaysia": ("malaysia", "malaysian"),
}


def normalized_words(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (value or "").lower()).strip()


def infer_label_context(text: str, cdlm_products: Optional[List[str]] = None,
                        cdlm_countries: Optional[List[str]] = None) -> Dict:
    """Read country, label type, drawing product and model from a guide PDF.

    A drawing title such as ``AV LBL 867247 Mexico box`` is more reliable than
    an inferred model: the drawing product (867247) is a CDLM column and the
    word ``box`` determines which CDLM location row is applicable.
    """
    normalized = normalized_words(text)
    country_hint = None
    country_evidence = None
    for country, aliases in COUNTRY_ALIASES.items():
        matched_alias = next((alias for alias in aliases
                              if re.search(rf"(?<![a-z]){re.escape(alias)}(?![a-z])", normalized)), None)
        if matched_alias:
            country_hint = country
            country_evidence = matched_alias
            break
    if not country_hint:
        for country in cdlm_countries or []:
            candidate = normalized_words(str(country))
            if candidate and re.search(rf"(?<![a-z0-9]){re.escape(candidate)}(?![a-z0-9])", normalized):
                country_hint = str(country)
                country_evidence = str(country)
                break
    model_match = re.search(r"(?:modelo|model)\s*:\s*(FM\s*-?\s*\d{1,3})\b", text or "", flags=re.IGNORECASE)
    if not model_match:
        model_match = re.search(r"\b(FM\s*-?\s*\d{1,3})\b", text or "", flags=re.IGNORECASE)
    model_hint = re.sub(r"\s|-", "", model_match.group(1).upper()) if model_match else None
    label_type_hint = None
    if re.search(r"\bbox(?:\s+label)?\b", normalized):
        label_type_hint = "box label"
    elif re.search(r"\bproduct\s+label\b", normalized):
        label_type_hint = "product label"
    product_hint = None
    for product in sorted(cdlm_products or [], key=lambda value: len(str(value)), reverse=True):
        value = str(product).strip()
        if value and re.search(rf"(?<![a-z0-9]){re.escape(value.lower())}(?![a-z0-9])", normalized):
            product_hint = value
            break
    return {
        "country_hint": country_hint,
        "model_hint": model_hint,
        "product_hint": product_hint,
        "label_type_hint": label_type_hint,
        "country_evidence": country_evidence,
        "country_source": "reference PDF text" if country_hint else None,
    }


def suggest_cdlm_selection(matrix: Dict, country_hint: Optional[str], model_hint: Optional[str],
                           product_hint: Optional[str] = None,
                           label_type_hint: Optional[str] = None) -> Dict:
    """Return a CDLM selection only when the country/model match is unique."""
    result = {"country": None, "product": None, "model": model_hint,
              "label_type": label_type_hint, "matched": False, "reason": None}
    if not country_hint:
        result["reason"] = "No unambiguous country was found on the reference PDF."
        return result
    hint_words = normalized_words(country_hint)
    country = next((value for value in matrix.get("countries", [])
                    if hint_words == normalized_words(value) or hint_words in COUNTRY_ALIASES.get(value, ())), None)
    if not country:
        result["reason"] = f"The drawing country '{country_hint}' is not present in the CDLM workbook."
        return result
    result["country"] = country
    location_hint = normalized_words(label_type_hint)
    entries = [entry for entry in matrix.get("entries", [])
               if entry.get("country") == country
               and (not location_hint or location_hint in normalized_words(entry.get("location", "")))]
    if product_hint:
        direct = sorted({entry["product"] for entry in entries
                         if normalized_words(entry.get("product", "")) == normalized_words(product_hint)})
        if len(direct) == 1:
            result.update({"product": direct[0], "matched": True,
                           "reason": f"Drawing product {direct[0]} maps to the {label_type_hint or 'applicable'} CDLM row."})
            return result
    if not model_hint:
        result["reason"] = f"Detected {country} from the reference PDF text; no drawing product or model was found to auto-select a CDLM product."
        return result
    model = normalized_words(model_hint).replace(" ", "")
    entries = [entry for entry in entries
               if model in normalized_words(entry.get("text", "")).replace(" ", "")]
    products = sorted({entry["product"] for entry in entries})
    if len(products) != 1:
        result["reason"] = ("No CDLM product" if not products else "More than one CDLM product") + f" matches {model_hint} for {country}."
        return result
    result.update({"product": products[0], "matched": True,
                   "reason": f"Drawing model {model_hint} maps to CDLM product {products[0]}."})
    return result
_blank_label_cache = None


# ════════════════════════════════════════════════════════════
# SYMBOL ASSET LOADING
# ════════════════════════════════════════════════════════════

def load_symbol_as_image(path: str) -> Optional[np.ndarray]:
    """Load a PNG or SVG symbol as an RGBA numpy array, trimmed of padding."""
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext == '.svg':
            if HAS_CAIRO:
                png_bytes = cairosvg.svg2png(url=path, dpi=300)
                img = Image.open(io.BytesIO(png_bytes)).convert('RGBA')
            else:
                # Fallback: try rendering SVG with PyMuPDF
                if HAS_FITZ:
                    doc = fitz.open(path)
                    page = doc[0]
                    pix = page.get_pixmap(dpi=300)
                    img_bytes = pix.tobytes("png")
                    img = Image.open(io.BytesIO(img_bytes)).convert('RGBA')
                    doc.close()
                else:
                    log.warning(f"No SVG renderer for {path}")
                    return None
        else:
            img = Image.open(path).convert('RGBA')

        arr = np.array(img)
        # Trim transparent/white padding
        arr = trim_padding(arr)
        if arr.shape[0] < 3 or arr.shape[1] < 3:
            log.warning(f"Symbol too small after trim: {path}")
            return None
        return arr
    except Exception as e:
        log.warning(f"Cannot load symbol {path}: {e}")
        return None


def trim_padding(img_rgba: np.ndarray) -> np.ndarray:
    """Remove transparent or near-white padding from all sides."""
    alpha = img_rgba[:, :, 3]
    # Consider a pixel as content if alpha > 10 AND not pure white
    rgb = img_rgba[:, :, :3]
    is_white = np.all(rgb > 240, axis=2)
    content_mask = (alpha > 10) & ~is_white
    rows = np.any(content_mask, axis=1)
    cols = np.any(content_mask, axis=0)
    if not rows.any() or not cols.any():
        return img_rgba
    r0, r1 = np.where(rows)[0][[0, -1]]
    c0, c1 = np.where(cols)[0][[0, -1]]
    return img_rgba[r0:r1+1, c0:c1+1]


def make_match_mask(img_rgba: np.ndarray) -> np.ndarray:
    """Create a binary mask for template matching (content pixels = 255)."""
    alpha = img_rgba[:, :, 3]
    rgb = img_rgba[:, :, :3]
    gray = cv2.cvtColor(rgb, cv2.COLOR_RGB2GRAY)
    # Content = has alpha AND is dark enough (not white background)
    content = (alpha > 30) & (gray < 200)
    mask = np.zeros_like(gray)
    mask[content] = 255
    return mask


def encode_image_b64(img_rgba: np.ndarray) -> str:
    """Encode an RGBA numpy array to base64 PNG string."""
    img = Image.fromarray(img_rgba)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    return base64.b64encode(buf.getvalue()).decode()


def normalize_symbol_id(symbol_id: str) -> str:
    """Normalize numeric and alphanumeric controlled symbol identifiers."""
    return re.sub(r"[^A-Za-z0-9]", "", str(symbol_id)).upper()


def get_symbol_specification(symbol_id: str) -> Optional[Dict]:
    """Return approved ingested metadata for a symbol ID, if it exists."""
    global _specification_cache
    normalized_id = normalize_symbol_id(symbol_id)
    if _specification_cache is None:
        catalog_path = _SYMBOLS / "symbol_specifications.json"
        catalog = {}
        try:
            raw_catalog = json.loads(catalog_path.read_text(encoding="utf-8"))
            records = raw_catalog.get("symbols", raw_catalog) if isinstance(raw_catalog, dict) else raw_catalog
            for record in records:
                if not isinstance(record, dict) or not record.get("symbol_id"):
                    continue
                if record.get("approved", True):
                    catalog[normalize_symbol_id(record["symbol_id"])] = record
        except (OSError, ValueError, TypeError) as exc:
            log.warning("Unable to load ingested symbol catalog: %s", exc)
        # Keep the packaged controls available until an approved ingestion
        # catalog is supplied. Ingested records take precedence over defaults.
        _specification_cache = {**SYMBOL_SPECIFICATIONS, **catalog}
    specification = _specification_cache.get(normalized_id)
    return dict(specification) if specification else None


def attach_symbol_specifications(symbols: List[Dict]) -> List[Dict]:
    """Add controlled specification metadata to matched symbols at generation time."""
    enriched = []
    for symbol in symbols:
        item = dict(symbol)
        specification = get_symbol_specification(item.get("code", ""))
        if specification:
            item["specification"] = specification
        enriched.append(item)
    return enriched


def load_blank_label_catalog() -> Dict:
    """Load the reviewed blank-label stock catalog extracted from the workbook."""
    global _blank_label_cache
    if _blank_label_cache is not None:
        return _blank_label_cache
    try:
        data = json.loads((_SYMBOLS / "blank_label_catalog.json").read_text(encoding="utf-8"))
        labels = [item for item in data.get("labels", []) if item.get("part_number")]
        _blank_label_cache = {"labels": labels, "error": None}
    except (OSError, ValueError, TypeError) as error:
        log.warning("Blank-label catalog unavailable: %s", error)
        _blank_label_cache = {"labels": [], "error": "Blank-label catalog is unavailable."}
    return _blank_label_cache


def load_drawing_catalog() -> Dict:
    """Load reviewed placement rectangles for controlled blank-label drawings.

    The part number is the join key: a label guide declares its blank-label
    stock and this catalog identifies the drawing that defines that stock.
    Keeping the rectangle in drawing coordinates avoids guessing placement
    from the raster preview and preserves the controlled drawing scale.
    """
    global _drawing_catalog_cache
    if _drawing_catalog_cache is not None:
        return _drawing_catalog_cache
    try:
        raw = json.loads((_DRAWINGS / "drawing_catalog.json").read_text(encoding="utf-8"))
        drawings = [item for item in raw.get("drawings", [])
                    if item.get("part_number") and item.get("file") and item.get("target_rect_pt")]
        _drawing_catalog_cache = {"drawings": drawings, "error": None}
    except (OSError, ValueError, TypeError) as error:
        log.warning("Drawing catalog unavailable: %s", error)
        _drawing_catalog_cache = {"drawings": [], "error": "No reviewed drawing placements are available."}
    return _drawing_catalog_cache


def drawing_for_part_number(part_number: str) -> Optional[Dict]:
    normalized = re.sub(r"[^0-9]", "", str(part_number or ""))
    return next((item for item in load_drawing_catalog()["drawings"]
                 if re.sub(r"[^0-9]", "", str(item.get("part_number", ""))) == normalized), None)


def load_layout_manifests() -> Dict:
    """Load reviewed physical layout slots for controlled reference labels."""
    global _layout_manifest_cache
    if _layout_manifest_cache is not None:
        return _layout_manifest_cache
    try:
        raw = json.loads((_SYMBOLS / "label_layout_manifests.json").read_text(encoding="utf-8"))
        layouts = raw.get("layouts", {}) if isinstance(raw, dict) else {}
        _layout_manifest_cache = layouts if isinstance(layouts, dict) else {}
    except (OSError, ValueError, TypeError) as error:
        log.warning("Layout manifests unavailable: %s", error)
        _layout_manifest_cache = {}
    return _layout_manifest_cache


def apply_layout_manifest(label_id: str, matched_symbols: List[Dict], country_region: Optional[Dict],
                          w_mm: float, h_mm: float, symbol_assets: Optional[List[Dict]] = None) -> Optional[Dict]:
    """Apply a reviewed layout manifest without reflowing the reference design."""
    manifest = load_layout_manifests().get(label_id)
    if not manifest or manifest.get("status") != "reviewed":
        return None

    reviewed_region = manifest.get("country_text_region")
    if isinstance(reviewed_region, dict) and country_region is not None:
        country_region.clear()
        country_region.update(reviewed_region)
        country_region.setdefault("text", "")

    slots = manifest.get("symbols", {})
    # A reviewed guide is authoritative: it explicitly lists the controlled
    # artwork allowed on that label.  Do not let a weak visual match from an
    # expanded asset library introduce a symbol not present in the guide.
    matched_symbols[:] = [symbol for symbol in matched_symbols if str(symbol.get("code", "")) in slots]
    present_codes = {str(symbol.get("code", "")) for symbol in matched_symbols}
    assets_by_code = {str(asset.get("code", "")): asset for asset in (symbol_assets or [])}
    for code in slots:
        if code in present_codes:
            continue
        asset = assets_by_code.get(code)
        if asset is None:
            log.warning("Reviewed manifest %s references missing asset %s", label_id, code)
            continue
        # The manifest is a reviewed extraction of this reference design. It
        # is therefore more reliable than vision matching for the presence of
        # its declared artwork, including wordmarks with sparse glyph shapes.
        matched_symbols.append({
            "asset": asset["file"], "code": code,
            "x": 0.0, "y": 0.0, "w": 0.0, "h": 0.0,
            "source_x": None, "source_y": None, "source_w": None, "source_h": None,
            "confidence": 1.0, "match_source": "reviewed-layout-manifest",
        })
    for symbol in matched_symbols:
        slot = slots.get(str(symbol.get("code", "")))
        if not isinstance(slot, dict):
            continue
        if slot.get("x_mm") is not None:
            symbol["x"] = round(float(slot["x_mm"]) / w_mm, 4)
        if slot.get("y_mm") is not None:
            symbol["y"] = round(float(slot["y_mm"]) / h_mm, 4)
        if slot.get("width_mm") is not None:
            prior_w = float(symbol.get("w", 0))
            prior_h = float(symbol.get("h", 0))
            symbol["w"] = round(float(slot["width_mm"]) / w_mm, 4)
            if slot.get("height_mm") is not None:
                symbol["h"] = round(float(slot["height_mm"]) / h_mm, 4)
            elif prior_w > 0 and prior_h > 0:
                symbol["h"] = round(symbol["w"] * prior_h / prior_w, 4)
        symbol["layout_source"] = "reviewed-layout-manifest"

    return {
        "status": "reviewed",
        "reference_pdf": manifest.get("reference_pdf"),
        "label_size_mm": manifest.get("label_size_mm"),
        "blank_label": manifest.get("blank_label", {}),
    }


def parse_country_label_workbook(workbook) -> Dict:
    """Read country/product label text from an opened CDLM workbook.

    The workbook identifies product applicability with an ``x`` in a product
    column.  Only the literal Text-column value is returned; values beginning
    with ``see`` are document references, not printable label text.
    """
    empty = {"countries": [], "products": [], "entries": [], "error": None}
    try:
        worksheet = workbook["Country Label"]
        product_columns = {
            column: str(worksheet.cell(2, column).value).strip()
            for column in range(11, worksheet.max_column + 1)
            if worksheet.cell(2, column).value
        }
        entries = []
        for row_number, row in enumerate(worksheet.iter_rows(min_row=3, values_only=True), start=3):
            country = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            text = str(row[8]).strip() if len(row) > 8 and row[8] else ""
            location = str(row[9]).strip() if len(row) > 9 and row[9] else ""
            if not country or not text:
                continue
            for column, product in product_columns.items():
                cell = row[column - 1] if len(row) >= column else None
                if str(cell).strip().lower() != "x":
                    continue
                entries.append({
                    "country": country,
                    "product": product,
                    "text": text,
                    "location": location,
                    "is_reference": text.lower().startswith("see "),
                    "source_row": row_number,
                })
        return {
            "countries": sorted({entry["country"] for entry in entries}),
            "products": sorted(product_columns.values()),
            "entries": entries,
            "error": None,
        }
    except Exception as error:
        log.error(f"CDLM read error: {error}", exc_info=True)
        empty["error"] = "Unable to read the CDLM workbook."
    return empty


def load_country_label_matrix() -> Dict:
    """Read country/product label text from the bundled CDLM workbook."""
    global _cdlm_cache
    if _cdlm_cache is not None:
        return _cdlm_cache
    if not HAS_OPENPYXL:
        _cdlm_cache = {"countries": [], "products": [], "entries": [], "error": "The CDLM reader dependency is unavailable."}
        return _cdlm_cache
    files = sorted(_CDLM.glob("LS-200004_CDLM_Avalon_Family_Rev*.xlsx"), reverse=True)
    if not files:
        # Backwards-compatible fallback for the originally bundled workbook.
        files = sorted(_SYMBOLS.glob("LS-200004_CDLM_Avalon_Family_Rev*.xlsx"), reverse=True)
    if not files:
        _cdlm_cache = {"countries": [], "products": [], "entries": [], "error": "No CDLM workbook was found in the ingested data folders."}
        return _cdlm_cache
    source = next((f for f in files if " (" not in f.name), files[0])
    workbook = None
    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
        _cdlm_cache = parse_country_label_workbook(workbook)
    finally:
        if workbook:
            workbook.close()
    return _cdlm_cache


def scan_symbol_assets():
    """Find all PNG/SVG symbol files. Pre-encode to base64 for fast responses."""
    assets = []
    asset_roots = [_SYMBOL_ASSETS, _SYMBOLS]
    source_files = []
    for root in asset_roots:
        if not root.is_dir():
            continue
        source_files.extend(sorted(path for path in root.glob('*') if path.suffix.lower() in {'.png', '.svg', '.jpg', '.jpeg'}))
    if not source_files:
        log.warning("No symbol asset files found in %s or %s", _SYMBOL_ASSETS, _SYMBOLS)
        return assets
    loaded_codes = set()
    for source in source_files:
        f = source.name
        low = f.lower()
        code_m = re.match(r'([A-Za-z0-9]+)(?=[_-]|$)', f)
        code = normalize_symbol_id(code_m.group(1) if code_m else source.stem)
        # assets/ is authoritative; avoid loading the legacy top-level copy.
        if code in loaded_codes:
            continue
        path = str(source)
        img = load_symbol_as_image(path)
        if img is None:
            continue
        # Pre-encode to base64 so we never need to serialize numpy later
        b64 = encode_image_b64(img)
        assets.append({
            'code': code,
            'file': f,
            'path': path,
            'image': img,          # numpy array for template matching
            'image_b64': b64,      # pre-encoded for API responses
            'h': int(img.shape[0]),
            'w': int(img.shape[1]),
            'is_svg': low.endswith('.svg')
        })
        loaded_codes.add(code)
        log.info(f"  Loaded: {f} ({img.shape[1]}x{img.shape[0]}, b64={len(b64)} chars)")
    # Prefer SVG over PNG for same code
    seen_codes = {}
    deduped = []
    for a in assets:
        if a['code'] in seen_codes:
            if a['is_svg'] and not seen_codes[a['code']]['is_svg']:
                deduped = [x for x in deduped if x['code'] != a['code']]
                deduped.append(a)
                seen_codes[a['code']] = a
        else:
            deduped.append(a)
            seen_codes[a['code']] = a
    log.info(f"Total symbol assets loaded: {len(deduped)}")
    return deduped


# ════════════════════════════════════════════════════════════
# PDF ANALYSIS
# ════════════════════════════════════════════════════════════

def detect_label_boundary(page, rendered_gray, w_mm=85, h_mm=50):
    """
    Detect the inner label rectangle using PyMuPDF vector paths.
    Uses page.get_drawings() to find the rounded/rectangular path whose
    aspect ratio matches the known label dimensions (85/50 = 1.7).

    Returns (x, y, w, h) in rendered-pixel coords, or None if not found.
    """
    expected_aspect = max(w_mm, h_mm) / min(w_mm, h_mm)  # 1.7
    scale = RENDER_DPI / 72  # PDF points -> rendered pixels
    page_w_pt = page.rect.width
    page_h_pt = page.rect.height
    page_area_pt = page_w_pt * page_h_pt

    # Expected label size in PDF points
    exp_w_pt = w_mm / 25.4 * 72  # ~240 pt
    exp_h_pt = h_mm / 25.4 * 72  # ~141 pt

    candidates = []

    # Strategy 1: page.get_drawings() — vector paths
    try:
        drawings = page.get_drawings()
        for d in drawings:
            rect = d.get('rect')  # fitz.Rect bounding box of the path
            if rect is None:
                continue
            rx, ry, rx1, ry1 = rect
            rw = rx1 - rx
            rh = ry1 - ry
            if rw < 20 or rh < 20:
                continue  # skip tiny marks, arrows, dimension ticks
            area = rw * rh
            # Reject paths that span the full page (border)
            if area > page_area_pt * 0.7:
                continue
            # Reject very small paths (< 3% of page)
            if area < page_area_pt * 0.02:
                continue
            aspect = max(rw, rh) / min(rw, rh)
            candidates.append((rx, ry, rw, rh, area, aspect))
    except Exception as e:
        log.warning(f"get_drawings() failed: {e}")

    # Strategy 2: also try page rects from annotations or explicit rect paths
    # (some PDFs encode the label outline as a simple rect, not a drawing)
    # We already have candidates from drawings; if empty, check page.rect

    if not candidates:
        log.warning("No vector paths found; cannot detect label boundary")
        return None

    # Score candidates: prefer aspect ratio close to 1.7 AND size close to expected
    best = None
    best_score = float('inf')

    for (rx, ry, rw, rh, area, aspect) in candidates:
        # Aspect error (try both landscape and portrait)
        asp_err = min(abs(aspect - expected_aspect),
                      abs(aspect - 1.0 / expected_aspect))
        if asp_err > 0.4:  # more than 40% off -> skip
            continue

        # Dimensional error: how close to expected 240x141 pt?
        dim_err1 = (abs(rw - exp_w_pt) / exp_w_pt +
                    abs(rh - exp_h_pt) / exp_h_pt)
        dim_err2 = (abs(rw - exp_h_pt) / exp_h_pt +
                    abs(rh - exp_w_pt) / exp_w_pt)
        dim_err = min(dim_err1, dim_err2)

        # Combined score: weight aspect more heavily
        score = asp_err * 2.0 + dim_err * 1.0
        if score < best_score:
            best_score = score
            best = (rx, ry, rw, rh)

    if best is None:
        log.warning("No candidate matched label aspect ratio 1.7")
        return None

    rx, ry, rw, rh = best
    log.info(f"  Label rect (PDF pts): ({rx:.1f}, {ry:.1f}) {rw:.1f}x{rh:.1f} "
             f"(score={best_score:.3f})")

    # Convert PDF points -> rendered pixel coords
    px = int(rx * scale)
    py = int(ry * scale)
    pw = int(rw * scale)
    ph = int(rh * scale)
    return (px, py, pw, ph)


def extract_text_region(page, label_bounds_pt, matched_symbols):
    """
    Extract text spans inside the label, exclude those overlapping symbols,
    and return ONE union bounding box as {x, y, w, h} in normalized 0-1 coords.
    Returns None if no text found.
    """
    lx, ly, lw, lh = label_bounds_pt
    if lw < 1 or lh < 1:
        return None

    text_dict = page.get_text("dict")
    span_boxes = []

    for block in text_dict.get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            for span in line.get("spans", []):
                bbox = span.get("bbox", [0, 0, 0, 0])
                sx0, sy0, sx1, sy1 = bbox
                if not (sx0 >= lx - 1 and sx1 <= lx + lw + 1 and
                        sy0 >= ly - 1 and sy1 <= ly + lh + 1):
                    continue
                nx = max(0.0, (sx0 - lx) / lw)
                ny = max(0.0, (sy0 - ly) / lh)
                nw = min(1.0, (sx1 - sx0) / lw)
                nh = min(1.0, (sy1 - sy0) / lh)
                # Exclude spans overlapping matched symbols
                overlaps = False
                for s in matched_symbols:
                    ssx = float(s.get('x', 0))
                    ssy = float(s.get('y', 0))
                    ssw = float(s.get('w', 0))
                    ssh = float(s.get('h', 0))
                    if (nx < ssx + ssw + 0.03 and nx + nw > ssx - 0.03 and
                        ny < ssy + ssh + 0.03 and ny + nh > ssy - 0.03):
                        overlaps = True
                        break
                if not overlaps and nw > 0.001 and nh > 0.001:
                    span_boxes.append((nx, ny, nw, nh))

    if not span_boxes:
        return None

    # Union bounding box
    x0 = min(b[0] for b in span_boxes)
    y0 = min(b[1] for b in span_boxes)
    x1 = max(b[0] + b[2] for b in span_boxes)
    y1 = max(b[1] + b[3] for b in span_boxes)
    return {'x': float(x0), 'y': float(y0), 'w': float(x1 - x0), 'h': float(y1 - y0)}


def extract_text_elements(page, label_bounds_px, matched_symbols, include_symbol_text=True):
    """Return editable PDF text spans in normalized label coordinates.

    ``include_symbol_text`` keeps codes or other vector text printed inside a
    symbol available to the human editor, while the country-text region can
    still deliberately omit it.
    """
    bx, by, bw, bh = label_bounds_px
    scale = RENDER_DPI / 72
    lx, ly, lw, lh = bx / scale, by / scale, bw / scale, bh / scale
    elements = []

    for block in page.get_text("dict").get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            for span in line.get("spans", []):
                text = str(span.get("text", "")).strip()
                sx0, sy0, sx1, sy1 = span.get("bbox", [0, 0, 0, 0])
                if not text or sx1 <= lx or sx0 >= lx + lw or sy1 <= ly or sy0 >= ly + lh:
                    continue
                x = max(0.0, (sx0 - lx) / lw)
                y = max(0.0, (sy0 - ly) / lh)
                w = min(1.0 - x, (sx1 - sx0) / lw)
                h = min(1.0 - y, (sy1 - sy0) / lh)
                if w < 0.002 or h < 0.002:
                    continue
                if not include_symbol_text and any(
                    x < s['x'] + s['w'] and x + w > s['x'] and
                    y < s['y'] + s['h'] and y + h > s['y']
                    for s in matched_symbols
                ):
                    continue
                elements.append({
                    'text': text,
                    'x': round(x, 4), 'y': round(y, 4),
                    'w': round(w, 4), 'h': round(h, 4),
                    'font_size': round(float(span.get('size', 8)) * 25.4 / 72, 2),
                })
    return elements


def country_text_region(text_elements):
    """Merge the label's printable text into one editable country-information area."""
    if not text_elements:
        return None
    x0 = min(item['x'] for item in text_elements)
    y0 = min(item['y'] for item in text_elements)
    x1 = max(item['x'] + item['w'] for item in text_elements)
    y1 = max(item['y'] + item['h'] for item in text_elements)
    return {
        'x': round(x0, 4), 'y': round(y0, 4),
        'w': round(x1 - x0, 4), 'h': round(y1 - y0, 4),
        'font_size': round(min(item['font_size'] for item in text_elements), 2),
        'text': '\n'.join(item['text'] for item in text_elements),
    }


def is_branding_text(text: str) -> bool:
    """Keep a manufacturer wordmark in the reference artwork, not in country text."""
    normalized = re.sub(r"\s+", "", str(text or "")).upper()
    return normalized in {"PHILIPS"}


def detect_source_text_bottom(label_crop: np.ndarray, matched_symbols: List[Dict]) -> Optional[float]:
    """Estimate the lowest outlined/source text line above the first symbol."""
    approved = [
        symbol for symbol in matched_symbols
        if (specification := get_symbol_specification(symbol.get('code', '')))
        and not specification.get('metadata_only')
    ]
    if not approved or label_crop is None:
        return None
    first_symbol_y = min(float(symbol.get('source_y', symbol['y'])) for symbol in approved)
    h, w = label_crop.shape[:2]
    y0, y1 = int(h * 0.16), int(h * max(0.16, first_symbol_y - 0.02))
    x0, x1 = int(w * 0.05), int(w * 0.95)
    if y1 <= y0:
        return None
    gray = cv2.cvtColor(label_crop, cv2.COLOR_BGR2GRAY) if label_crop.ndim == 3 else label_crop
    dark = gray[y0:y1, x0:x1] < 100
    row_counts = np.count_nonzero(dark, axis=1)
    rows = np.where(row_counts >= max(8, int((x1 - x0) * 0.01)))[0]
    if not len(rows):
        return None
    return round((y0 + int(rows[-1]) + 1) / h, 4)


def fallback_country_text_region(matched_symbols: List[Dict], source_text_bottom: Optional[float] = None) -> Dict:
    """Reserve the controlled text zone for outlined reference-label PDFs.

    Some controlled drawings convert their lettering to vector outlines, so a
    PDF text extractor returns no spans.  The reference layout still gives a
    reliable structure: branding at the top, country text above the approved
    symbols.  Only symbols with an approved specification are considered here
    so an incidental graphic match cannot shrink the text box.
    """
    approved_symbols = [
        symbol for symbol in matched_symbols
        if (specification := get_symbol_specification(symbol.get('code', '')))
        and not specification.get('metadata_only')
    ]
    symbol_top = min((symbol['y'] for symbol in approved_symbols), default=0.76)
    # Thai labels begin with the country-specific importer line.  Brazilian
    # labels reserve the upper band for the PHILIPS wordmark, so their text
    # field begins below that brand area.
    is_thai_layout = any(symbol.get('code') == '100183' for symbol in approved_symbols)
    left = approved_symbols[0]['x'] if is_thai_layout and approved_symbols else 0.06
    top = 0.06 if is_thai_layout else 0.18
    bottom = max(top + 0.16, (source_text_bottom + 0.02) if source_text_bottom else symbol_top - 0.025)
    return {
        'x': round(left, 4),
        'y': top,
        'w': 0.88,
        'h': round(max(0.16, bottom - top), 4),
        'font_size': 1.6,
        'text': '',
        'detected_from': 'reference-layout',
    }


def thai_symbol_number_region(matched_symbols: List[Dict]) -> Optional[Dict]:
    """Return the editable notification-number area within a matched Thai FDA symbol."""
    thai = next((symbol for symbol in matched_symbols if symbol.get('code') == '100183'), None)
    if not thai:
        return None
    # The logo and outer frame stay untouched.  This is only the white text
    # bay inside the frame where the country-specific notification number sits.
    return {
        'x': round(float(thai['x']) + float(thai['w']) * 0.20, 4),
        'y': round(float(thai['y']) + float(thai['h']) * 0.22, 4),
        # Extend far enough to erase the entire previous number (including its
        # final digits), while remaining inside the right-hand sloping frame.
        'w': round(float(thai['w']) * 0.69, 4),
        'h': round(float(thai['h']) * 0.56, 4),
        'font_size': 1.7,
        'detected_from': 'thai-fda-symbol',
    }


def build_text_mask(page, label_bounds_px, render_dpi):
    """Build binary mask of text regions inside the label crop."""
    bx, by, bw, bh = label_bounds_px
    inv_scale = 72 / render_dpi
    lx_pt, ly_pt = bx * inv_scale, by * inv_scale
    lw_pt, lh_pt = bw * inv_scale, bh * inv_scale
    scale = render_dpi / 72
    mask = np.zeros((bh, bw), dtype=np.uint8)
    text_dict = page.get_text("dict")
    for block in text_dict.get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            for span in line.get("spans", []):
                sx0, sy0, sx1, sy1 = span.get("bbox", [0, 0, 0, 0])
                if sx1 < lx_pt or sx0 > lx_pt + lw_pt:
                    continue
                if sy1 < ly_pt or sy0 > ly_pt + lh_pt:
                    continue
                px0 = max(0, int((sx0 - lx_pt) * scale))
                py0 = max(0, int((sy0 - ly_pt) * scale))
                px1 = min(bw, int((sx1 - lx_pt) * scale))
                py1 = min(bh, int((sy1 - ly_pt) * scale))
                if px1 > px0 and py1 > py0:
                    mask[py0:py1, px0:px1] = 255
    return mask


def find_graphic_components(label_crop, text_mask, min_area_pct=0.002):
    """Find connected graphic blobs after removing text from label."""
    h, w = label_crop.shape[:2]
    min_area = int(h * w * min_area_pct)
    _, binary = cv2.threshold(label_crop, 180, 255, cv2.THRESH_BINARY_INV)
    binary[text_mask > 0] = 0
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
    binary = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel)
    num_labels, labels, stats, _ = cv2.connectedComponentsWithStats(binary, 8)
    components = []
    for i in range(1, num_labels):
        cx = int(stats[i, cv2.CC_STAT_LEFT])
        cy = int(stats[i, cv2.CC_STAT_TOP])
        cw = int(stats[i, cv2.CC_STAT_WIDTH])
        ch = int(stats[i, cv2.CC_STAT_HEIGHT])
        area = int(stats[i, cv2.CC_STAT_AREA])
        if area < min_area:
            continue
        if cw > w * 0.95 and ch > h * 0.95:
            continue
        comp_mask = (labels[cy:cy+ch, cx:cx+cw] == i).astype(np.uint8) * 255
        components.append({'x': cx, 'y': cy, 'w': cw, 'h': ch,
                           'area': area, 'mask': comp_mask})
    components.sort(key=lambda c: c['area'], reverse=True)
    return components


def match_asset_to_component(symbol_asset, component):
    """Score how well an asset matches a component (IoU + aspect)."""
    comp_mask = component['mask']
    ch, cw = comp_mask.shape
    sym_rgba = symbol_asset['image']
    sym_match_mask = make_match_mask(sym_rgba)
    if np.sum(sym_match_mask > 0) == 0:
        return 0.0
    mask_resized = cv2.resize(sym_match_mask, (cw, ch), interpolation=cv2.INTER_AREA)
    _, sym_bin = cv2.threshold(mask_resized, 50, 255, cv2.THRESH_BINARY)
    sym_fg = sym_bin > 0
    comp_fg = comp_mask > 0
    intersection = int(np.sum(sym_fg & comp_fg))
    union = int(np.sum(sym_fg | comp_fg))
    if union == 0:
        return 0.0
    iou = intersection / union
    sym_h, sym_w = sym_rgba.shape[:2]
    sym_asp = sym_w / max(sym_h, 1)
    comp_asp = cw / max(ch, 1)
    asp_sim = min(sym_asp, comp_asp) / max(sym_asp, comp_asp)
    return iou * (0.5 + 0.5 * asp_sim)


def component_match_pipeline(page, label_crop, label_bounds_px, symbol_assets,
                             w_mm=85, h_mm=50):
    """Component-based matching: mask text, find blobs, match assets.
    Keep detected component dimensions unless a controlled specification overrides them."""
    bx, by, bw, bh = label_bounds_px
    text_mask = build_text_mask(page, label_bounds_px, RENDER_DPI)
    components = find_graphic_components(label_crop, text_mask)
    log.info(f"  {len(components)} graphic components found")
    for i, c in enumerate(components[:10]):
        log.info(f"    comp[{i}]: ({c['x']},{c['y']}) {c['w']}x{c['h']} area={c['area']}")

    matched = []
    unmatched = []
    reference_graphics = []
    # Score every asset/component pairing first, then accept the strongest
    # mutually unique pairs. The old asset-by-asset greedy loop made whichever
    # PNG was alphabetically first claim a component, leaving the correct
    # symbol unavailable and producing extra marks as the library expanded.
    scores = [[match_asset_to_component(asset, component) for component in components]
              for asset in symbol_assets]
    candidates = []
    for asset_index, asset_scores in enumerate(scores):
        for component_index, score in enumerate(asset_scores):
            other_scores = sorted((row[component_index] for row in scores), reverse=True)
            runner_up = other_scores[1] if len(other_scores) > 1 else 0.0
            if score >= MATCH_THRESHOLD and score - runner_up >= MATCH_UNIQUENESS_MARGIN:
                candidates.append((score, asset_index, component_index))
    candidates.sort(reverse=True)
    used_assets, used_components = set(), set()
    accepted = {}
    for score, asset_index, component_index in candidates:
        if asset_index in used_assets or component_index in used_components:
            continue
        accepted[asset_index] = (component_index, score)
        used_assets.add(asset_index)
        used_components.add(component_index)

    for asset_index, asset in enumerate(symbol_assets):
        selected = accepted.get(asset_index)
        best_score = max(scores[asset_index], default=0.0)
        if selected:
            best_idx, best_score = selected
            comp = components[best_idx]
            # Start with the size detected on the source label.
            norm_w = float(comp['w']) / bw
            norm_h = float(comp['h']) / bh
            sym_w_mm = norm_w * w_mm
            sym_h_mm = norm_h * h_mm

            # Apply only a specification that explicitly controls size.
            specification = get_symbol_specification(asset['code'])
            if specification and not specification.get('metadata_only') and specification.get('required_width_mm'):
                sym_w_mm = specification['required_width_mm']
                sym_h_mm = sym_w_mm / (asset['w'] / max(asset['h'], 1))
                norm_w = sym_w_mm / w_mm
                norm_h = sym_h_mm / h_mm
            elif specification and not specification.get('metadata_only') and specification.get('minimum_size_mm'):
                min_size_mm = specification['minimum_size_mm']
                detected_max_mm = max(sym_w_mm, sym_h_mm)
                if detected_max_mm < min_size_mm:
                    scale = min_size_mm / max(detected_max_mm, 0.001)
                    sym_w_mm *= scale
                    sym_h_mm *= scale
                    norm_w = sym_w_mm / w_mm
                    norm_h = sym_h_mm / h_mm

            # Center on the detected component, while preserving a 1 mm clear
            # margin on every side of the blank label.
            cx = (comp['x'] + comp['w'] / 2) / bw
            cy = (comp['y'] + comp['h'] / 2) / bh
            margin_x = LABEL_EDGE_MARGIN_MM / w_mm
            margin_y = LABEL_EDGE_MARGIN_MM / h_mm
            min_x, min_y = margin_x, margin_y
            max_x = max(min_x, 1.0 - margin_x - norm_w)
            max_y = max(min_y, 1.0 - margin_y - norm_h)
            norm_x = min(max(min_x, cx - norm_w / 2), max_x)
            norm_y = min(max(min_y, cy - norm_h / 2), max_y)
            matched.append({
                'asset': asset['file'], 'code': asset['code'],
                'x': round(norm_x, 4),
                'y': round(norm_y, 4),
                'w': round(norm_w, 4),
                'h': round(norm_h, 4),
                # Preserve the original detected artwork box. A controlled
                # symbol can later be reflowed without leaving a second copy
                # of the reference artwork behind.
                'source_x': round(float(comp['x']) / bw, 4),
                'source_y': round(float(comp['y']) / bh, 4),
                'source_w': round(float(comp['w']) / bw, 4),
                'source_h': round(float(comp['h']) / bh, 4),
                'confidence': round(best_score, 4)
            })
            log.info(f"  MATCH {asset['file']} -> comp[{best_idx}] "
                     f"score={best_score:.4f} size={sym_w_mm:.1f}x{sym_h_mm:.1f}mm")
        else:
            unmatched.append({
                'asset': asset['file'],
                'reason': 'not present in this label' if best_score < 0.05
                          else f'below threshold ({best_score:.3f})',
                'confidence': round(best_score, 4)
            })
            log.info(f"  SKIP {asset['file']}: best={best_score:.4f}")

    # Preserve only genuinely source-only pictograms which are not yet
    # represented by an ingested symbol PNG.  Segmentation can split one
    # approved symbol into several adjacent components; preserving one of
    # those fragments would paint a second, partial copy over the approved
    # symbol.  Therefore a candidate must be both clearly unrelated to every
    # library asset and at least 1 mm clear of every matched component.
    clear_x_px = max(1, round(bw / max(w_mm, 1) * LABEL_EDGE_MARGIN_MM))
    clear_y_px = max(1, round(bh / max(h_mm, 1) * LABEL_EDGE_MARGIN_MM))

    def intersects_matched_component(component):
        left = component['x'] - clear_x_px
        top = component['y'] - clear_y_px
        right = component['x'] + component['w'] + clear_x_px
        bottom = component['y'] + component['h'] + clear_y_px
        for matched_index in used_components:
            matched_component = components[matched_index]
            if (left < matched_component['x'] + matched_component['w'] + clear_x_px
                    and right > matched_component['x'] - clear_x_px
                    and top < matched_component['y'] + matched_component['h'] + clear_y_px
                    and bottom > matched_component['y'] - clear_y_px):
                return True
        return False

    for component_index, component in enumerate(components):
        if component_index in used_components:
            continue
        norm_w = float(component['w']) / bw
        norm_h = float(component['h']) / bh
        if not (0.004 <= norm_w <= 0.16 and 0.004 <= norm_h <= 0.16):
            continue
        best_asset_score = max((row[component_index] for row in scores), default=0.0)
        if best_asset_score >= 0.05 or intersects_matched_component(component):
            log.info(f"  SKIP reference fragment comp[{component_index}]: "
                     f"asset-score={best_asset_score:.3f} or adjacent to matched symbol")
            continue
        pad = 2
        x0 = max(0, component['x'] - pad)
        y0 = max(0, component['y'] - pad)
        x1 = min(bw, component['x'] + component['w'] + pad)
        y1 = min(bh, component['y'] + component['h'] + pad)
        encoded, graphic_png = cv2.imencode('.png', label_crop[y0:y1, x0:x1])
        if not encoded:
            continue
        reference_graphics.append({
            'x': round(float(x0) / bw, 4), 'y': round(float(y0) / bh, 4),
            'w': round(float(x1 - x0) / bw, 4), 'h': round(float(y1 - y0) / bh, 4),
            'image': base64.b64encode(graphic_png.tobytes()).decode(),
            'source': 'unmatched-reference-graphic',
        })

    # Text region: normalized union of text mask extent
    text_coords = np.where(text_mask > 0)
    text_region = None
    if len(text_coords[0]) > 0:
        ty0 = int(text_coords[0].min())
        ty1 = int(text_coords[0].max())
        tx0 = int(text_coords[1].min())
        tx1 = int(text_coords[1].max())
        text_region = {
            'x': float(tx0) / bw, 'y': float(ty0) / bh,
            'w': float(tx1 - tx0) / bw, 'h': float(ty1 - ty0) / bh
        }

    return matched, unmatched, len(components), text_region, reference_graphics


def template_match_symbol(rendered_gray, symbol_asset, label_bounds_px):
    """LEGACY fallback: multi-scale template matching. Not used in new pipeline."""
    lx, ly, lw, lh = label_bounds_px
    # Crop rendered image to label area
    label_region = rendered_gray[ly:ly+lh, lx:lx+lw]
    if label_region.size == 0:
        return None

    # Create grayscale template from symbol
    sym_rgba = symbol_asset['image']
    sym_gray = cv2.cvtColor(sym_rgba[:, :, :3], cv2.COLOR_RGB2GRAY)
    sym_mask = make_match_mask(sym_rgba)

    # Diagnostic: verify mask has foreground pixels
    mask_fg_pixels = int(np.sum(sym_mask > 0))
    if mask_fg_pixels == 0:
        log.warning(f"  {symbol_asset['file']}: mask has ZERO fg pixels, skipping")
        return None

    # Invert: in PDF, symbols are dark on light; template should match dark content
    _, sym_bin = cv2.threshold(sym_gray, 128, 255, cv2.THRESH_BINARY_INV)
    # Also threshold the label region
    _, label_bin = cv2.threshold(label_region, 180, 255, cv2.THRESH_BINARY_INV)

    best_match = None
    best_val = 0

    # Multi-scale matching
    template_h, template_w = sym_bin.shape
    label_h_px, label_w_px = label_bin.shape

    # Scale range: symbol could be 5% to 60% of label width
    min_w = int(label_w_px * 0.03)
    max_w = int(label_w_px * 0.70)

    # Generate scales
    scales = []
    current = min_w
    while current <= max_w:
        scale_factor = current / template_w
        scaled_h = int(template_h * scale_factor)
        if scaled_h > 5 and scaled_h < label_h_px - 2 and current < label_w_px - 2:
            scales.append((current, scaled_h))
        current = int(current * 1.15)  # 15% increments

    for (tw, th) in scales:
        resized_template = cv2.resize(sym_bin, (tw, th), interpolation=cv2.INTER_AREA)
        resized_mask = cv2.resize(sym_mask, (tw, th), interpolation=cv2.INTER_NEAREST)

        # Skip if template is larger than search region
        if tw >= label_w_px or th >= label_h_px:
            continue

        try:
            result = cv2.matchTemplate(label_bin, resized_template,
                                       cv2.TM_CCORR_NORMED, mask=resized_mask)
            _, max_val, _, max_loc = cv2.minMaxLoc(result)

            # Reject inf/nan (mask normalization error)
            if max_val != max_val or max_val == float('inf'):
                continue
            if max_val > best_val and max_val <= 1.0:
                best_val = max_val
                best_match = {
                    'x': float(max_loc[0]) / label_w_px,
                    'y': float(max_loc[1]) / label_h_px,
                    'w': float(tw) / label_w_px,
                    'h': float(th) / label_h_px,
                    'confidence': float(max_val)
                }
        except cv2.error:
            continue

    if best_match and best_match['confidence'] >= MATCH_THRESHOLD:
        return best_match
    elif best_match:
        log.info(f"  Low confidence ({best_match['confidence']:.3f}) for "
                 f"{symbol_asset['file']} — not placed")
    return None


# ════════════════════════════════════════════════════════════
# MAIN PIPELINE
# ════════════════════════════════════════════════════════════

MIN_SYMBOL_SIZE = 0.02  # reject matches < 2% of label dimension


def process_pdf_label(pdf_path, symbol_assets, output_dpi=600, label_id: Optional[str] = None):
    """
    Crop-first pipeline:
    1. Rasterize full PDF page
    2. Extract text to get label dimensions (mm)
    3. Detect inner label rectangle via vector paths
    4. CROP rendered image to label boundary only
    5. Template-match symbols within cropped label image
    6. Validate: reject <2% matches, require >=2 symbols
    7. Compute text region (one union box excluding symbols)
    """
    if not HAS_FITZ or not HAS_CV2:
        return None

    fname = os.path.basename(pdf_path)
    log.info(f"Processing: {fname}")

    try:
        doc = fitz.open(pdf_path)
    except Exception as e:
        log.error(f"Cannot open {fname}: {e}")
        return None

    page = doc[0]

    # 1. Extract text first to get dimensions
    full_text = page.get_text()
    w_mm, h_mm = 85, 50
    m = re.search(r'\(?(\d+)\s*mm\s*[xX\u00d7]\s*(\d+)\s*mm\)?', full_text)
    if m:
        d1, d2 = int(m.group(1)), int(m.group(2))
        w_mm, h_mm = max(d1, d2), min(d1, d2)
    log.info(f"  Dimensions: {w_mm}x{h_mm}mm")

    # 2. Rasterize full page
    mat = fitz.Matrix(RENDER_DPI / 72, RENDER_DPI / 72)
    pix = page.get_pixmap(matrix=mat)
    img_bytes = pix.tobytes("png")
    full_rendered = cv2.imdecode(np.frombuffer(img_bytes, np.uint8), cv2.IMREAD_GRAYSCALE)
    img_h, img_w = full_rendered.shape
    log.info(f"  Full page: {img_w}x{img_h}px")

    # 3. Detect inner label boundary using vector paths
    label_bounds_px = detect_label_boundary(page, full_rendered, w_mm, h_mm)
    if label_bounds_px is None:
        log.error(f"  FAILED: no label boundary in {fname}")
        doc.close()
        return {
            'id': fname.replace('.pdf', ''),
            'title': fname.replace('.pdf', ''),
            'error': 'No label boundary detected via vector paths',
            'w_mm': w_mm, 'h_mm': h_mm,
            'symbols': [], 'text_region': None,
            'debug': {'error': 'no vector path matched aspect 1.7'}
        }

    bx, by, bw, bh = label_bounds_px
    log.info(f"  Label rect: ({bx},{by}) {bw}x{bh}px")

    # 4. CROP rendered image to label only
    bx = max(0, min(bx, img_w - 1))
    by = max(0, min(by, img_h - 1))
    bw = min(bw, img_w - bx)
    bh = min(bh, img_h - by)
    if bw < 20 or bh < 20:
        log.error(f"  Crop too small: {bw}x{bh}")
        doc.close()
        return {
            'id': fname.replace('.pdf', ''),
            'title': fname.replace('.pdf', ''),
            'error': f'Label crop too small ({bw}x{bh}px)',
            'w_mm': w_mm, 'h_mm': h_mm,
            'symbols': [], 'text_region': None,
            'debug': {'label_bounds_px': [bx, by, bw, bh]}
        }

    label_crop = full_rendered[by:by+bh, bx:bx+bw]
    log.info(f"  Cropped label: {label_crop.shape[1]}x{label_crop.shape[0]}px")

    # Keep the approved drawing artwork as the preview base. This preserves
    # printed text and layout rather than rebuilding a blank label from only
    # detected symbols.
    encoded, label_png = cv2.imencode('.png', label_crop)
    label_image_b64 = base64.b64encode(label_png.tobytes()).decode() if encoded else None

    # 5. Component-based matching (text masked, graphic blobs isolated, IoU scored)
    matched_symbols, failed_symbols, n_components, text_region, reference_graphics = \
        component_match_pipeline(page, label_crop, (bx, by, bw, bh), symbol_assets,
                                 w_mm=w_mm, h_mm=h_mm)
    text_elements = extract_text_elements(page, (bx, by, bw, bh), matched_symbols)
    country_elements = extract_text_elements(
        page, (bx, by, bw, bh), matched_symbols, include_symbol_text=False
    )
    printable_country_elements = [
        element for element in country_elements if not is_branding_text(element['text'])
    ]
    editable_country_region = country_text_region(
        printable_country_elements
    )
    # Prefer the actual extracted text baseline. Some controlled PDFs outline
    # their lettering, in which case inspect the rendered reference artwork.
    source_text_bottom = max(
        (float(element['y']) + float(element['h']) for element in printable_country_elements),
        default=None,
    )
    if source_text_bottom is None:
        source_text_bottom = detect_source_text_bottom(label_crop, matched_symbols)
    if editable_country_region is None:
        editable_country_region = fallback_country_text_region(
            matched_symbols, source_text_bottom
        )

    # Fallback only for templates that do not yet have a reviewed layout
    # manifest. Controlled existing labels use their manifest immediately
    # after this compatibility path.
    controlled_symbols = [
        symbol for symbol in matched_symbols
        if (specification := get_symbol_specification(symbol.get('code', '')))
        and not specification.get('metadata_only') and symbol.get('code') != '100183'
    ]
    if len(controlled_symbols) == 1 and source_text_bottom is not None:
        symbol = controlled_symbols[0]
        symbol['y'] = round(min(
            1.0 - LABEL_EDGE_MARGIN_MM / h_mm - symbol['h'],
            max(LABEL_EDGE_MARGIN_MM / h_mm, source_text_bottom + LABEL_EDGE_MARGIN_MM / h_mm)
        ), 4)
    template_id = label_id or fname.replace('.pdf', '')
    layout_manifest = apply_layout_manifest(
        template_id, matched_symbols, editable_country_region, w_mm, h_mm, symbol_assets
    )
    if layout_manifest:
        log.info("  Applied reviewed layout manifest for %s", fname)
    editable_thai_symbol_region = thai_symbol_number_region(matched_symbols)
    n_matched = len(matched_symbols)
    log.info(f"  Result: {n_matched} matched, {len(failed_symbols)} skipped, "
             f"{n_components} components")

    # Title
    title = fname.replace('.pdf', '')
    tm_match = re.search(r'TITLE\n(.+)', full_text)
    if tm_match:
        title = tm_match.group(1).strip()

    doc.close()
    cdlm_matrix = load_country_label_matrix()
    cdlm_products = cdlm_matrix.get("products", [])

    return {
        'id': template_id,
        'title': title,
        'w_mm': w_mm, 'h_mm': h_mm,
        'symbols': matched_symbols,
        'text_region': text_region,
        'text_elements': text_elements,
        'country_text_region': editable_country_region,
        'thai_symbol_region': editable_thai_symbol_region,
        'reference_graphics': reference_graphics,
        'layout_manifest': layout_manifest,
        'inferred_context': infer_label_context(full_text, cdlm_products, cdlm_matrix.get("countries", [])),
        'label_image': label_image_b64,
        'debug': {
            'render_dpi': RENDER_DPI,
            'page_size_px': f"{img_w}x{img_h}",
            'label_bounds_px': [bx, by, bw, bh],
            'label_crop_px': f"{bw}x{bh}",
            'pipeline': 'component-match',
            'assets_tested': len(symbol_assets),
            'assets_matched': n_matched,
            'graphic_components_found': n_components,
            'unmatched_assets': failed_symbols,
            'reference_graphics_preserved': len(reference_graphics),
            'has_text_region': text_region is not None
        }
    }


def _dead_code_start():  # pragma: no cover
    """Everything below until _dead_code_end was the OLD matching logic."""

    for asset in symbol_assets:
        match = template_match_symbol(label_crop, asset, crop_bounds)
        if match:
            # Reject too-small matches (<2% of label)
            if match['w'] < MIN_SYMBOL_SIZE or match['h'] < MIN_SYMBOL_SIZE:
                reason = f"too small ({match['w']:.4f}x{match['h']:.4f})"
                log.warning(f"  REJECT {asset['file']}: {reason}")
                failed_symbols.append({'asset': asset['file'],
                                       'reason': reason,
                                       'confidence': match['confidence']})
                continue
            # Reject outside bounds
            if (match['x'] < -0.01 or match['y'] < -0.01 or
                match['x'] + match['w'] > 1.05 or
                match['y'] + match['h'] > 1.05):
                reason = f"outside label (x={match['x']:.3f} y={match['y']:.3f})"
                log.warning(f"  REJECT {asset['file']}: {reason}")
                failed_symbols.append({'asset': asset['file'],
                                       'reason': reason,
                                       'confidence': match['confidence']})
                continue
            matched_symbols.append({
                'asset': asset['file'], 'code': asset['code'],
                'x': match['x'], 'y': match['y'],
                'w': match['w'], 'h': match['h'],
                'confidence': match['confidence']
            })
            log.info(f"  OK {asset['file']}: "
                     f"({match['x']:.3f},{match['y']:.3f}) "
                     f"{match['w']:.3f}x{match['h']:.3f} "
                     f"conf={match['confidence']:.3f}")
        else:
            failed_symbols.append({'asset': asset['file'],
                                   'reason': 'not present in this label',
                                   'confidence': 0})
            log.info(f"  SKIP {asset['file']}: not in this PDF")

    # 6. Each PDF uses its own subset of symbols — do NOT require all
    n_matched = len(matched_symbols)
    n_not_present = len([f for f in failed_symbols if f['confidence'] == 0])
    n_rejected = len([f for f in failed_symbols if f['confidence'] > 0])
    log.info(f"  Result: {n_matched} matched, {n_not_present} not in this PDF, "
             f"{n_rejected} rejected (size/bounds)")

    # Only fail if ZERO symbols matched AND label has visible content
    if n_matched == 0:
        dark_pix = int(np.sum(label_crop < 128))
        total_pix = label_crop.shape[0] * label_crop.shape[1]
        has_content = dark_pix > total_pix * 0.01
        if has_content:
            log.error(f"  FAIL: 0 matched but label has content ({dark_pix/total_pix*100:.1f}% dark)")
            doc.close()
            return {
                'id': fname.replace('.pdf', ''),
                'title': fname.replace('.pdf', ''),
                'error': f'No symbols matched (0/{len(symbol_assets)}). Label has visible content but no asset correlated.',
                'w_mm': w_mm, 'h_mm': h_mm,
                'symbols': [],
                'failed_symbols': failed_symbols,
                'text_region': None,
                'debug': {
                    'label_bounds_px': [bx, by, bw, bh],
                    'label_crop_px': f"{bw}x{bh}",
                    'dark_pixel_pct': f"{dark_pix/total_pix*100:.1f}%",
                    'failed_symbols': failed_symbols
                }
            }
        else:
            log.warning(f"  Label crop appears blank")
            doc.close()
            return None

    # 7. Text region (union box excluding symbol areas)
    scale = 72 / RENDER_DPI
    label_bounds_pt = (bx * scale, by * scale, bw * scale, bh * scale)
    text_region = extract_text_region(page, label_bounds_pt, matched_symbols)
    log.info(f"  Text region: {text_region}")

    # Title
    title = fname.replace('.pdf', '')
    tm_match = re.search(r'TITLE\n(.+)', full_text)
    if tm_match:
        title = tm_match.group(1).strip()

    doc.close()

    return {
        'id': fname.replace('.pdf', ''),
        'title': title,
        'w_mm': w_mm, 'h_mm': h_mm,
        'symbols': matched_symbols,
        'text_region': text_region,
        'debug': {
            'render_dpi': RENDER_DPI,
            'page_size_px': f"{img_w}x{img_h}",
            'label_bounds_px': [bx, by, bw, bh],
            'label_crop_px': f"{bw}x{bh}",
            'assets_tested': len(symbol_assets),
            'assets_matched': len(matched_symbols),
            'failed_symbols': failed_symbols,
            'has_text_region': text_region is not None
        }
    }


def _dead_code_end():  # pragma: no cover
    pass


# ════════════════════════════════════════════════════════════
# CATALOG
# ════════════════════════════════════════════════════════════

def load_catalog():
    global _cache, _cache_t
    now = time.time()
    if _cache and (now - _cache_t) < 300:
        return _cache

    assets = scan_symbol_assets()
    log.info(f"Symbol assets: {[a['file'] for a in assets]}")

    labels = []
    label_files = []
    for root in [_LABEL_GUIDES, _SYMBOLS]:
        if root.is_dir():
            label_files.extend(sorted(path for path in root.glob('*.pdf') if 'mart' in path.name.lower()))
    seen_guides = set()
    for guide in label_files:
        if guide.name in seen_guides:
            continue
        seen_guides.add(guide.name)
        result = process_pdf_label(str(guide), assets)
        if result:
            labels.append(result)

    _cache = {'labels': labels, 'assets': assets}
    _cache_t = now
    return _cache


# ════════════════════════════════════════════════════════════
# API
# ════════════════════════════════════════════════════════════

@app.get("/", response_class=HTMLResponse)
async def home():
    p = _APP / "index.html"
    return HTMLResponse(p.read_text()) if p.exists() else HTMLResponse("<h1>B300 v6</h1>")


@app.get("/api/catalog")
async def api_catalog():
    c = load_catalog()
    return {
        "products": [{
            'product_code': lab['id'].split('-M-')[0],
            'product_desc': lab['title'],
            'label_size': f"{lab['h_mm']} X {lab['w_mm']} mm",
            'sheet_name': lab['id'],
            'symbol_count': len(lab['symbols']),
            'inferred_context': lab.get('inferred_context', {})
        } for lab in c['labels']],
        "count": len(c['labels'])
    }


def cdlm_response(matrix: Dict, country: Optional[str], product: Optional[str],
                  label_type: Optional[str] = None) -> Dict:
    """Build the public CDLM response for bundled or temporary workbooks."""
    response = {"countries": matrix["countries"], "products": matrix["products"]}
    if country and product:
        requested_type = normalized_words(label_type)
        response["entries"] = [
            entry for entry in matrix["entries"]
            if entry["country"] == country
            and entry["product"] == product
            and (requested_type in normalized_words(entry["location"])
                 if requested_type else "product label" in entry["location"].lower())
        ]
    return response


@app.get("/api/country-labels/suggest")
async def api_country_label_suggestion(country_hint: Optional[str] = None, model_hint: Optional[str] = None,
                                       product_hint: Optional[str] = None, label_type_hint: Optional[str] = None,
                                       session_id: Optional[str] = None):
    session = _uploaded_cdlm_sessions.get(session_id or "")
    if session and time.time() - session["created_at"] <= 30 * 60:
        matrix = session["matrix"]
    elif session_id:
        return JSONResponse(content={"error": "The uploaded CDLM workbook session has expired. Upload it again."}, status_code=410)
    else:
        matrix = load_country_label_matrix()
    if matrix["error"]:
        return JSONResponse(content={"error": matrix["error"]}, status_code=503)
    return suggest_cdlm_selection(matrix, country_hint, model_hint, product_hint, label_type_hint)


@app.get("/api/country-labels")
async def api_country_labels(country: Optional[str] = None, product: Optional[str] = None,
                             label_type: Optional[str] = None,
                             session_id: Optional[str] = None):
    """Expose CDLM options and safe product-label text for the selected pair."""
    session = _uploaded_cdlm_sessions.get(session_id or "")
    if session and time.time() - session["created_at"] <= 30 * 60:
        matrix = session["matrix"]
    elif session_id:
        return JSONResponse(content={"error": "The uploaded CDLM workbook session has expired. Upload it again."}, status_code=410)
    else:
        matrix = load_country_label_matrix()
    if matrix["error"]:
        return JSONResponse(content={"error": matrix["error"]}, status_code=503)
    return cdlm_response(matrix, country, product, label_type)


@app.get("/api/blank-labels")
async def api_blank_labels():
    """Expose reviewed blank-label size and material records for validation."""
    catalog = load_blank_label_catalog()
    if catalog["error"]:
        return JSONResponse(content={"error": catalog["error"]}, status_code=503)
    return catalog


@app.get("/api/drawings")
async def api_drawings():
    """Expose only the reviewed blank-label drawings available for output."""
    return load_drawing_catalog()


@app.post("/api/country-labels/upload")
async def api_upload_country_labels(file: UploadFile = File(...)):
    """Parse an uploaded CDLM XLSX in memory for this browser session only."""
    if not HAS_OPENPYXL:
        raise HTTPException(503, "The CDLM reader dependency is unavailable.")
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(400, "Upload an .xlsx CDLM workbook.")
    payload = await file.read()
    if not payload or len(payload) > MAX_UPLOAD_BYTES:
        raise HTTPException(400, "The workbook must be between 1 byte and 25 MB.")
    workbook = None
    try:
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        matrix = parse_country_label_workbook(workbook)
    except Exception as error:
        log.warning("Uploaded CDLM read error: %s", error)
        raise HTTPException(422, "Unable to read this CDLM workbook. It must contain a 'Country Label' sheet.")
    finally:
        if workbook:
            workbook.close()
    if matrix["error"] or not matrix["entries"]:
        raise HTTPException(422, matrix["error"] or "No country/product label records were found in this workbook.")
    now = time.time()
    for key, session in list(_uploaded_cdlm_sessions.items()):
        if now - session["created_at"] > 30 * 60:
            del _uploaded_cdlm_sessions[key]
    session_id = uuid.uuid4().hex
    _uploaded_cdlm_sessions[session_id] = {"created_at": now, "matrix": matrix}
    return {
        **cdlm_response(matrix, None, None),
        "session_id": session_id,
        "filename": file.filename,
        "temporary": True,
    }


@app.get("/api/symbols/{symbol_id}")
async def api_symbol_specification(symbol_id: str):
    """Return the controlled specification for one symbol."""
    specification = get_symbol_specification(symbol_id)
    if not specification:
        raise HTTPException(404, f"No specification found for symbol {symbol_id}")
    return specification


@app.post("/api/export-png")
async def api_export_png(request: LabelExportRequest):
    """Build a physical-size PNG server-side; avoids browser SVG limitations."""
    try:
        width = round(request.w_mm / 25.4 * request.dpi)
        height = round(request.h_mm / 25.4 * request.dpi)
        blank_label = (request.layout_manifest or {}).get("blank_label", {})
        clean_render = bool(blank_label.get("clean_render"))
        if clean_render:
            image = Image.new("RGB", (width, height), "white")
        else:
            source = base64.b64decode(request.label_image, validate=True)
            image = Image.open(io.BytesIO(source)).convert("RGB").resize(
                (width, height), Image.Resampling.LANCZOS
            )
        native_template = image.copy()
        draw = ImageDraw.Draw(image)

        if clean_render:
            radius = max(0, round(float(blank_label.get("corner_radius_mm", 0)) / 25.4 * request.dpi))
            border = max(1, round(0.25 / 25.4 * request.dpi))
            draw.rounded_rectangle(
                (border // 2, border // 2, width - 1 - border // 2, height - 1 - border // 2),
                radius=radius,
                outline="black",
                width=border,
            )

        # First replace the source country text.  Controlled symbols are
        # painted afterwards: their approved artwork must remain whole even
        # when a deliberately generous editable text region crosses the
        # source symbol's original location.
        if request.country_region:
            render_country_html(draw, request.country_region, request.country_html, request.dpi, width, height)

        for graphic in request.reference_graphics:
            raw_graphic = graphic.get("image", "")
            if not raw_graphic:
                continue
            try:
                x = round(float(graphic.get("x", 0)) * width)
                y = round(float(graphic.get("y", 0)) * height)
                graphic_w = max(1, round(float(graphic.get("w", 0)) * width))
                graphic_h = max(1, round(float(graphic.get("h", 0)) * height))
                graphic_image = Image.open(io.BytesIO(base64.b64decode(raw_graphic, validate=True))).convert("RGB")
                image.paste(graphic_image.resize((graphic_w, graphic_h), Image.Resampling.LANCZOS), (x, y))
            except Exception:
                log.warning("Skipping unreadable reference graphic during PNG export")

        for symbol in request.symbols:
            # 100183's generic PNG has no Thai notification number. Preserve
            # the complete, numbered symbol already present in the reference
            # label image rather than replacing it with an incomplete asset.
            if not clean_render and (not symbol.get("specification") or symbol.get("code") == "100183"
                    or symbol["specification"].get("metadata_only")):
                continue
            raw_asset = request.symbol_images.get(symbol.get("asset", ""))
            if not raw_asset:
                continue
            x = round(float(symbol.get("x", 0)) * width)
            y = round(float(symbol.get("y", 0)) * height)
            symbol_w = max(1, round(float(symbol.get("w", 0)) * width))
            symbol_h = max(1, round(float(symbol.get("h", 0)) * height))
            asset = Image.open(io.BytesIO(base64.b64decode(raw_asset, validate=True))).convert("RGBA")
            asset = asset.resize((symbol_w, symbol_h), Image.Resampling.LANCZOS)
            if not clean_render:
                # Clear the original detected artwork (with a small bleed
                # margin) before drawing at its controlled/reflowed position.
                source_x = round(float(symbol.get("source_x", symbol.get("x", 0))) * width)
                source_y = round(float(symbol.get("source_y", symbol.get("y", 0))) * height)
                source_w = max(1, round(float(symbol.get("source_w", symbol.get("w", 0))) * width))
                source_h = max(1, round(float(symbol.get("source_h", symbol.get("h", 0))) * height))
                mask_pad = max(2, round(request.dpi / 40))
                draw.rectangle((source_x - mask_pad, source_y - mask_pad,
                                source_x + source_w + mask_pad, source_y + source_h + mask_pad), fill="white")
            draw.rectangle((x, y, x + symbol_w, y + symbol_h), fill="white")
            image.paste(asset, (x, y), asset)

        thai_symbol = next((symbol for symbol in request.symbols if symbol.get("code") == "100183"), None)
        thai_y = None
        if not clean_render and thai_symbol and request.thai_symbol_y is not None:
            original_x = round(float(thai_symbol.get("x", 0)) * width)
            original_y = round(float(thai_symbol.get("y", 0)) * height)
            symbol_w = max(1, round(float(thai_symbol.get("w", 0)) * width))
            symbol_h = max(1, round(float(thai_symbol.get("h", 0)) * height))
            thai_y = round(float(request.thai_symbol_y) * height)
            native_symbol = native_template.crop((original_x, original_y, original_x + symbol_w, original_y + symbol_h))
            # The reference crop has a white background, so masking the old
            # position preserves the blank label before the native artwork is moved.
            draw.rectangle((original_x, original_y, original_x + symbol_w, original_y + symbol_h), fill="white")
            image.paste(native_symbol, (original_x, thai_y))
        if request.thai_symbol_region:
            thai_number_region = dict(request.thai_symbol_region)
            if thai_symbol and thai_y is not None:
                thai_number_region["y"] = float(thai_number_region["y"]) + (
                    float(request.thai_symbol_y) - float(thai_symbol.get("y", 0))
                )
            render_thai_symbol_text(
                draw, thai_number_region, request.thai_symbol_text,
                request.dpi, width, height
            )

        output = io.BytesIO()
        image.save(output, format="PNG", dpi=(request.dpi, request.dpi), optimize=True)
        output.seek(0)
        safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "-", request.label_id or "label")
        return StreamingResponse(
            output,
            media_type="image/png",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}-{request.dpi}dpi.png"'},
        )
    except Exception as exc:
        log.error("PNG export failed: %s", exc, exc_info=True)
        raise HTTPException(422, "Unable to create the PNG export.")


@app.post("/api/export-drawing")
async def api_export_drawing(request: DrawingExportRequest):
    """Place the completed label into its matching controlled drawing PDF.

    The source drawing remains unchanged.  The returned PDF is a new copy with
    the generated physical-size label placed into its reviewed target rectangle.
    """
    if not HAS_FITZ:
        raise HTTPException(503, "Drawing PDF export is not available in this app runtime.")
    drawing = drawing_for_part_number(request.drawing_part_number)
    if not drawing:
        raise HTTPException(422, "No reviewed drawing is available for the selected blank label stock.")
    drawing_path = _DRAWINGS / Path(str(drawing["file"])).name
    if not drawing_path.is_file():
        raise HTTPException(422, "The matched drawing file is not available to this app.")
    target = drawing.get("target_rect_pt", [])
    if not isinstance(target, list) or len(target) != 4:
        raise HTTPException(422, "The matched drawing has no valid label placement rectangle.")
    try:
        png_response = await api_export_png(request)
        png_bytes = bytearray()
        async for chunk in png_response.body_iterator:
            png_bytes.extend(chunk)
        if not png_bytes:
            raise ValueError("Label PNG renderer returned no data")

        document = fitz.open(drawing_path)
        page_index = int(drawing.get("page", 0))
        if page_index < 0 or page_index >= len(document):
            raise ValueError("Drawing target page is outside the source document")
        rect = fitz.Rect(*[float(value) for value in target])
        page = document[page_index]
        rotation = int(drawing.get("rotation", 0))
        if rotation not in {0, 90, 180, 270}:
            raise ValueError("Drawing placement rotation must be a right angle")
        page.insert_image(rect, stream=bytes(png_bytes), keep_proportion=False,
                          overlay=True, rotate=rotation)
        output = document.tobytes(garbage=4, deflate=True)
        document.close()
        safe_label = re.sub(r"[^A-Za-z0-9_.-]+", "-", request.label_id or "label")
        part = re.sub(r"[^0-9]", "", request.drawing_part_number)
        return Response(
            content=output,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{safe_label}-on-{part}-drawing.pdf"'},
        )
    except HTTPException:
        raise
    except Exception as exc:
        log.error("Drawing PDF export failed: %s", exc, exc_info=True)
        raise HTTPException(422, "Unable to create the drawing PDF export.")


@app.post("/api/interpret-edit")
async def api_interpret_edit(request: EditRequest):
    """Interpret a human editing request with ai_extract on the app warehouse.

    This endpoint deliberately returns a proposed, structured operation only.
    It never writes the drawing, PDF, specification, or regulatory data.
    """
    warehouse_id = os.getenv("DATABRICKS_WAREHOUSE_ID")
    if not warehouse_id:
        return JSONResponse(
            content={"error": "AI editing is not configured for this app."},
            status_code=503,
        )

    command = request.command.strip()
    schema_literal = json.dumps(AI_EDIT_SCHEMA, separators=(",", ":")).replace("'", "''")
    command_literal = "'" + command.replace("'", "''") + "'"
    statement = (
        "SELECT to_json(ai_extract("
        f"{command_literal}, '{schema_literal}', map('version','2.1')):response) AS edit"
    )

    try:
        from databricks.sdk import WorkspaceClient

        result = WorkspaceClient().statement_execution.execute_statement(
            warehouse_id=warehouse_id,
            statement=statement,
            wait_timeout="30s",
        )
        state = str(getattr(getattr(result, "status", None), "state", ""))
        if "SUCCEEDED" not in state:
            message = getattr(getattr(result, "status", None), "error", None) or "AI request did not complete."
            return JSONResponse(content={"error": str(message)}, status_code=502)
        rows = getattr(getattr(result, "result", None), "data_array", None) or []
        if not rows or not rows[0]:
            return JSONResponse(content={"error": "AI returned no proposed edit."}, status_code=502)
        proposed_edit = unwrap_ai_extract_values(json.loads(rows[0][0]))
        return {"proposal": proposed_edit, "review_required": True}
    except Exception as exc:
        log.error("AI edit interpretation failed: %s", exc, exc_info=True)
        return JSONResponse(
            content={"error": "Unable to interpret this edit request right now."},
            status_code=502,
        )


def build_generation_response(lab, assets):
    """Build the shared API payload for a stored or temporarily uploaded PDF."""
    if lab.get('error'):
        return JSONResponse(content={
            "error": lab['error'],
            "label_id": lab['id'],
            "failed_symbols": sanitize(lab.get('failed_symbols', [])),
            "matched_symbols": sanitize(lab.get('symbols', [])),
            "debug": sanitize(lab.get('debug', {}))
        }, status_code=422)

    sym_images = {}
    for sym in lab['symbols']:
        asset = next((a for a in assets if a['file'] == sym['asset']), None)
        if asset and asset.get('image_b64'):
            sym_images[sym['asset']] = asset['image_b64']

    symbols = attach_symbol_specifications(lab['symbols'])
    return JSONResponse(content={
        "label_id": lab['id'],
        "title": lab['title'],
        "label_size": f"{lab['h_mm']} X {lab['w_mm']} mm",
        "w_mm": int(lab['w_mm']),
        "h_mm": int(lab['h_mm']),
        "symbols": sanitize(symbols),
        "symbol_specifications": [s['specification'] for s in symbols if 'specification' in s],
        "text_region": sanitize(lab.get('text_region')),
        "text_elements": sanitize(lab.get('text_elements', [])),
        "country_text_region": sanitize(lab.get('country_text_region')),
        "country_text_source": "reference PDF text" if (lab.get('country_text_region') or {}).get('text') else None,
        "thai_symbol_region": sanitize(lab.get('thai_symbol_region')),
        "reference_graphics": sanitize(lab.get('reference_graphics', [])),
        "layout_manifest": sanitize(lab.get('layout_manifest')),
        "inferred_context": sanitize(lab.get('inferred_context', {})),
        "label_image": lab.get('label_image'),
        "symbol_images": sym_images,
        "symbols_placed": len(lab['symbols']),
        "convention": "template-matched",
        "debug": sanitize(lab.get('debug', {}))
    })


@app.get("/api/generate/{label_id}")
async def api_generate(label_id: str, dpi: int = 600):
    """Generate label: returns normalized symbol positions + base64 images."""
    try:
        c = load_catalog()
        lab = next((l for l in c['labels'] if l['id'] == label_id), None)
        if not lab:
            avail = [l['id'] for l in c['labels']]
            return JSONResponse(content={"error": f"Not found: {label_id}", "available": avail}, status_code=404)

        return build_generation_response(lab, c['assets'])
    except Exception as e:
        log.error(f"Generate error for {label_id}: {e}", exc_info=True)
        return JSONResponse(content={"error": str(e)}, status_code=500)


@app.post("/api/generate-upload")
async def api_generate_upload(file: UploadFile = File(...), dpi: int = 600):
    """Process one user-uploaded PDF from temporary storage, then delete it."""
    filename = Path(file.filename or "uploaded-label.pdf").name
    if not filename.lower().endswith(".pdf"):
        raise HTTPException(400, "Please upload a PDF file.")

    temp_path = None
    try:
        with tempfile.NamedTemporaryFile(prefix="label-upload-", suffix=".pdf", delete=False) as tmp:
            temp_path = tmp.name
            total_bytes = 0
            while chunk := await file.read(1024 * 1024):
                total_bytes += len(chunk)
                if total_bytes > MAX_UPLOAD_BYTES:
                    raise HTTPException(413, "PDF exceeds the 25 MB upload limit.")
                tmp.write(chunk)

        with open(temp_path, "rb") as uploaded:
            if uploaded.read(5) != b"%PDF-":
                raise HTTPException(400, "The uploaded file is not a valid PDF.")

        assets = scan_symbol_assets()
        lab = process_pdf_label(temp_path, assets, label_id=Path(filename).stem)
        if not lab:
            return JSONResponse(content={
                "error": "No usable label boundary or content was found in this PDF."
            }, status_code=422)

        # Keep the original drawing name in the response; the temporary path is never exposed.
        lab['title'] = Path(filename).stem
        return build_generation_response(lab, assets)
    except HTTPException:
        raise
    except Exception as e:
        log.error(f"Generate error for uploaded PDF {filename}: {e}", exc_info=True)
        return JSONResponse(content={"error": "Unable to process this PDF."}, status_code=422)
    finally:
        await file.close()
        if temp_path:
            try:
                os.unlink(temp_path)
            except FileNotFoundError:
                pass


@app.get("/api/generate/{label_id}/image")
async def api_generate_image(label_id: str, dpi: int = 600, debug_overlay: bool = False):
    """Render label as a single PNG image with symbols composited."""
    c = load_catalog()
    lab = next((l for l in c['labels'] if l['id'] == label_id), None)
    if not lab:
        raise HTTPException(404, f"Not found: {label_id}")

    w_px = int(lab['w_mm'] / 25.4 * dpi)
    h_px = int(lab['h_mm'] / 25.4 * dpi)

    # White canvas
    canvas = Image.new('RGB', (w_px, h_px), 'white')

    # Place symbols
    for sym in lab['symbols']:
        asset = next((a for a in c['assets'] if a['file'] == sym['asset']), None)
        if not asset:
            continue
        sx = int(sym['x'] * w_px)
        sy = int(sym['y'] * h_px)
        sw = int(sym['w'] * w_px)
        sh = int(sym['h'] * h_px)
        if sw < 1 or sh < 1:
            continue
        sym_img = Image.fromarray(asset['image']).resize((sw, sh), Image.LANCZOS)
        # Paste with alpha
        if sym_img.mode == 'RGBA':
            canvas.paste(sym_img, (sx, sy), sym_img)
        else:
            canvas.paste(sym_img, (sx, sy))

    # Debug overlay
    if debug_overlay:
        from PIL import ImageDraw, ImageFont
        draw = ImageDraw.Draw(canvas)
        # Label border
        draw.rectangle([2, 2, w_px-3, h_px-3], outline='blue', width=2)
        # Symbol boxes
        for sym in lab['symbols']:
            sx = int(sym['x'] * w_px)
            sy = int(sym['y'] * h_px)
            sw = int(sym['w'] * w_px)
            sh = int(sym['h'] * h_px)
            draw.rectangle([sx, sy, sx+sw, sy+sh], outline='red', width=1)
            draw.text((sx, sy-10),
                      f"{sym['asset']} ({sym['confidence']:.2f})",
                      fill='red')
        # Text region box
        tr = lab.get('text_region')
        if tr:
            tx = int(float(tr['x']) * w_px)
            ty = int(float(tr['y']) * h_px)
            tw = int(float(tr['w']) * w_px)
            th = int(float(tr['h']) * h_px)
            draw.rectangle([tx, ty, tx+tw, ty+th], outline='green', width=1)

    buf = io.BytesIO()
    canvas.save(buf, format='PNG', dpi=(dpi, dpi))
    buf.seek(0)

    if debug_overlay:
        img_b64 = base64.b64encode(buf.getvalue()).decode()
        return {"image": f"data:image/png;base64,{img_b64}",
                "product_code": lab['id'].split('-M-')[0],
                "product_desc": lab['title'],
                "label_size": f"{lab['h_mm']} X {lab['w_mm']} mm",
                "symbols_placed": len(lab['symbols']),
                "convention": "template-matched"}

    return Response(content=buf.getvalue(), media_type="image/png")


@app.get("/api/download/{label_id}")
async def api_download(label_id: str, dpi: int = 600):
    resp = await api_generate_image(label_id, dpi, debug_overlay=False)
    if isinstance(resp, Response):
        resp.headers["Content-Disposition"] = f'attachment; filename="{label_id}.png"'
        return resp
    raise HTTPException(500, "Unexpected response type")


@app.get("/api/health")
async def api_health():
    try:
        c = load_catalog()
        data = {
            "status": "ok",
            "dependencies": {
                "PyMuPDF": HAS_FITZ, "OpenCV": HAS_CV2, "CairoSVG": HAS_CAIRO
            },
            "symbols_dir": str(_SYMBOLS),
            "dir_exists": os.path.isdir(str(_SYMBOLS)),
            "all_files": os.listdir(str(_SYMBOLS)) if os.path.isdir(str(_SYMBOLS)) else [],
            "assets_loaded": [{'file': a['file'], 'size': f"{a['w']}x{a['h']}"}
                              for a in c['assets']],
            "labels": [{
                'id': l['id'],
                'symbols_matched': len(l['symbols']),
                'matched': [{'asset': s['asset'],
                             'confidence': float(s['confidence']),
                             'pos': f"({float(s['x']):.2f},{float(s['y']):.2f})",
                             'size': f"{float(s['w']):.2f}x{float(s['h']):.2f}"}
                            for s in l['symbols']],
                'error': l.get('error'),
                'has_text_region': l.get('text_region') is not None,
                'debug': l.get('debug', {})
            } for l in c['labels']]
        }
        return JSONResponse(content=json.loads(json.dumps(data, cls=NumpyEncoder)))
    except Exception as e:
        log.error(f"Health check error: {e}", exc_info=True)
        return JSONResponse(content={"status": "error", "message": str(e)}, status_code=500)


@app.on_event("startup")
async def startup():
    log.info("=== B300 Label Generator v6 (PyMuPDF + OpenCV) ===")
    log.info(f"PyMuPDF: {HAS_FITZ}, OpenCV: {HAS_CV2}, CairoSVG: {HAS_CAIRO}")
    log.info(f"Symbols dir: {_SYMBOLS} (exists={os.path.isdir(str(_SYMBOLS))})")
    if os.path.isdir(str(_SYMBOLS)):
        log.info(f"Contents: {os.listdir(str(_SYMBOLS))}")
    # Pre-warm cache
    try:
        load_catalog()
    except Exception as e:
        log.error(f"Startup cache error: {e}", exc_info=True)


@app.get("/api/test")
async def api_test():
    """Simple test: returns OK if app is running."""
    return JSONResponse(content={"ok": True, "msg": "App is alive"})


@app.get("/api/test_generate/{label_id}")
async def api_test_generate(label_id: str):
    """Debug generate: returns step-by-step trace to find crash point."""
    steps = []
    try:
        steps.append("1. Loading catalog...")
        c = load_catalog()
        steps.append(f"2. Labels: {[l['id'] for l in c['labels']]}")
        steps.append(f"3. Assets: {[a['file'] for a in c['assets']]}")

        lab = next((l for l in c['labels'] if l['id'] == label_id), None)
        if not lab:
            steps.append(f"4. FAIL: '{label_id}' not in labels")
            return JSONResponse(content={"steps": steps})

        steps.append(f"4. Found label: {lab['id']}")
        steps.append(f"5. Symbols: {len(lab['symbols'])}")
        steps.append(f"6. Text spans: {len(lab['text_spans'])}")

        # Check symbol types
        for i, sym in enumerate(lab['symbols']):
            steps.append(f"7.{i} sym keys={list(sym.keys())} types={{k: type(v).__name__ for k,v in sym.items()}}")

        # Try building sym_images
        steps.append("8. Building sym_images...")
        sym_images = {}
        for sym in lab['symbols']:
            asset = next((a for a in c['assets'] if a['file'] == sym['asset']), None)
            if asset and asset.get('image_b64'):
                sym_images[sym['asset']] = asset['image_b64'][:50] + "..."
                steps.append(f"   OK: {sym['asset']} ({len(asset['image_b64'])} chars)")
            else:
                steps.append(f"   MISS: {sym['asset']} (asset={asset is not None})")

        # Try sanitize
        steps.append("9. Sanitizing symbols...")
        clean_syms = sanitize(lab['symbols'])
        steps.append(f"10. Sanitized type: {type(clean_syms).__name__}")
        if clean_syms:
            steps.append(f"11. First sym: {clean_syms[0]}")

        steps.append("12. Building result dict...")
        result = {
            "label_id": str(lab['id']),
            "w_mm": int(lab['w_mm']),
            "h_mm": int(lab['h_mm']),
            "symbols": clean_syms,
            "symbol_images": sym_images,
            "symbols_placed": len(lab['symbols'])
        }
        steps.append("13. Returning JSONResponse...")
        return JSONResponse(content={"steps": steps, "result_keys": list(result.keys())})
    except Exception as e:
        steps.append(f"CRASH: {type(e).__name__}: {e}")
        import traceback
        steps.append(traceback.format_exc())
        return JSONResponse(content={"steps": steps}, status_code=500)
